from django.shortcuts import render
from django.shortcuts import redirect
from django.utils import timezone
from django.contrib.auth import login, authenticate
from django.core.mail import send_mail
from django.core.mail import EmailMessage
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from openpyxl import Workbook

from .models import Worker
from .models import Usuario
from .models import Post
from .forms import UsuarioForm
from .forms import ContactForm
from .forms import WorkForm
import os

	
# Create your views here.

def thanks(request):
    return render(request, 'blog/mensaje.html')

def bloga(request):
    posts = Post.objects.filter(fecha__lte=timezone.now()).order_by('-fecha')
    return render(request, 'blog/indexa.html', {'posts': posts})

def posta(request,postid):

    post = Post.objects.get(id=postid)
    return render(request, 'blog/post.html', {'post': post})

def about(request):
    return render(request, 'blog/about.html')
	
def trabaja(request):
    if request.method == 'POST':

        form = WorkForm(request.POST)
        if form.is_valid() and ('myfile' in request.FILES):
            print('Con adjunto')
            BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            data = form.cleaned_data
            myfile = request.FILES['myfile']
            fs = FileSystemStorage()
            filename = fs.save(myfile.name, myfile)
            uploaded_file_url = fs.url(filename)
            print(uploaded_file_url)
          # send_mail(
          #     'Bolsa de trabajo', 
          #     ' El Trabajador: ' + data['name'] + ' Con email: ' + data['email'] + ' Mensaje: ' + data['message'] + ' Telefono: ' + data['phone'] 
          #     + ' dni: ' + data['dni'] + ' lugar de residencia: ' + data['ciudad'] + ' edad: ' + data['edad'] + ' nacionalidad: ' + data['nacionalidad']
          #     + ' formacion general: ' + data['formacion']  + ' formacion específica: ' + data['especifica'] + ' cuidados: ' + data['cuidados'] + ' remunerados: ' + data['remunerados'] + ' duracion: ' + data['duracion']
          #     + ' dependencia: ' + data['dependencia'] + ' parcial: ' + data['parcial'] + ' acompanamiento: ' + data['acompanamiento'] + ' interno: ' + data['interno']
          #     + ' media_jornada: ' + data['media_jornada'] + ' entre_semana: ' + data['entre_semana'] + ' fines_semana: ' + data['fines_semana'] + ' vehiculo: ' + data['vehiculo']
          #     + ' desplazamiento: ' + data['desplazamiento'] + ' solo_localidad: ' + data['solo_localidad'] + ' provincia: ' + data['provincia'],
          #     data['email'], #FROM
          #     ['info@manospararespirar.com'],
          #     fail_silently=False,
          # )
            print(data['email'])

            wb = Workbook()
            #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
            ws = wb.active
            #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
            ws['B1'] = 'Trabajador'
            #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
            ws.merge_cells('B1:E1')
            #Creamos los encabezados desde la celda B3 hasta la E3
            ws['B3'] = 'APELLIDO 1'
            ws['C3'] = 'APELLIDO 2'
            ws['D3'] = 'NOMBRE'
            ws['E3'] = 'EDAD' 
            ws['F3'] = 'SEXO'
            ws['G3'] = 'PAIS'
            ws['H3'] = 'FORMACION'
            ws['I3'] = 'ESPECIFICA' 
            ws['J3'] = 'EXPERIENCIA'
            ws['K3'] = 'HORARIOS'
            ws['L3'] = 'RESIDENCIA'
            ws['M3'] = 'IDIOMAS' 
            ws['N3'] = 'VEHICULO'
            ws['O3'] = 'EMAIL'
            ws['P3'] = 'TELEFONO'
            ws['Q3'] = 'APTO'
            ws['R3'] = 'OBSERVACIONES'
            ws['S3'] = 'DNI' 
            ws['T3'] = 'MESES EXPERENCIA'
            ws['U3'] = 'RADIO DESPLAZAMIENTO'
            ws['V3'] = 'SOLO EN LOCALIDAD'
            ws['W3'] = 'EN TODA LA PROVINCIA'
            ws['X3'] = 'COMENTARIOS TRABAJADOR'  				
            cont=4
            #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
            #for persona in personas:
            ws.cell(row=cont,column=2).value = data['surname']
            ws.cell(row=cont,column=4).value = data['name'] 
            ws.cell(row=cont,column=5).value = data['edad']
            ws.cell(row=cont,column=7).value = data['nacionalidad']
            ws.cell(row=cont,column=8).value = data['formacion'] 
            ws.cell(row=cont,column=9).value = data['especifica']
            ws.cell(row=cont,column=11).value = 'Disponibilidad media jornada: ' + data['media_jornada'] + ' Disponibilidad entre semana: ' + data['entre_semana'] + 'Disponibilidad fines de semana: ' + data['fines_semana'] 
            ws.cell(row=cont,column=12).value = data['ciudad']
            ws.cell(row=cont,column=14).value = data['vehiculo'] 
            ws.cell(row=cont,column=15).value = data['email']
            ws.cell(row=cont,column=16).value = data['phone'] 
            ws.cell(row=cont,column=19).value = data['dni']
            ws.cell(row=cont,column=20).value = data['duracion']
            ws.cell(row=cont,column=21).value = data['desplazamiento']
            ws.cell(row=cont,column=22).value = data['solo_localidad']
            ws.cell(row=cont,column=23).value = data['provincia']
            ws.cell(row=cont,column=24).value = data['message']
            #cont = cont + 1
            #Establecemos el nombre del archivo
            nombre_archivo ="DatosTrabajador" + data['dni'] +".xlsx"
            wb.save(nombre_archivo)

            msg = EmailMessage('Bolsa de trabajo', 'Contacto recibido, se adjunta el curriculum y la información en el excel', data['email'], ['info@manospararespirar.com'])
            msg.content_subtype = "html"  
            msg.attach_file(BASE_DIR + uploaded_file_url)
            msg.attach_file(nombre_archivo)
            msg.send()
			
            return render(request, 'blog/mensaje.html', {})
        else:
            if form.is_valid():
                print('Sin adjunto')
                data = form.cleaned_data
              # send_mail(
              # 'Bolsa de trabajo',
			  # ' El Trabajador: ' + data['name'] + ' Con email: ' + data['email'] + ' Mensaje: ' + data['message'] + ' Telefono: ' + data['phone'] 
              # + ' dni: ' + data['dni'] + ' lugar de residencia: ' + data['ciudad'] + ' edad: ' + data['edad'] + ' nacionalidad: ' + data['nacionalidad']
              # + ' formacion general: ' + data['formacion']  + ' formacion específica: ' + data['especifica'] + ' cuidados: ' + data['cuidados'] + ' remunerados: ' + data['remunerados'] + ' duracion: ' + data['duracion']
              # + ' dependencia: ' + data['dependencia'] + ' parcial: ' + data['parcial'] + ' acompanamiento: ' + data['acompanamiento'] + ' interno: ' + data['interno']
              # + ' media_jornada: ' + data['media_jornada'] + ' entre_semana: ' + data['entre_semana'] + ' fines_semana: ' + data['fines_semana'] + ' vehiculo: ' + data['vehiculo']
              # + ' desplazamiento: ' + data['desplazamiento'] + ' solo_localidad: ' + data['solo_localidad'] + ' provincia: ' + data['provincia'],
              # data['email'], #FROM
              # ['info@manospararespirar.com'],
              # fail_silently=False,
              #  )

                wb = Workbook()
                #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
                ws = wb.active
                #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
                ws['B1'] = 'Trabajador'
                #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
                ws.merge_cells('B1:E1')
                #Creamos los encabezados desde la celda B3 hasta la E3
                ws['B3'] = 'APELLIDO 1'
                ws['C3'] = 'APELLIDO 2'
                ws['D3'] = 'NOMBRE'
                ws['E3'] = 'EDAD' 
                ws['F3'] = 'SEXO'
                ws['G3'] = 'PAIS'
                ws['H3'] = 'FORMACION'
                ws['I3'] = 'ESPECIFICA' 
                ws['J3'] = 'EXPERIENCIA'
                ws['K3'] = 'HORARIOS'
                ws['L3'] = 'RESIDENCIA'
                ws['M3'] = 'IDIOMAS' 
                ws['N3'] = 'VEHICULO'
                ws['O3'] = 'EMAIL'
                ws['P3'] = 'TELEFONO'
                ws['Q3'] = 'APTO'
                ws['R3'] = 'OBSERVACIONES'
                ws['S3'] = 'DNI' 
                ws['T3'] = 'MESES EXPERENCIA'
                ws['U3'] = 'RADIO DESPLAZAMIENTO'
                ws['V3'] = 'SOLO EN LOCALIDAD'
                ws['W3'] = 'EN TODA LA PROVINCIA'
                ws['X3'] = 'COMENTARIOS TRABAJADOR'  				
                cont=4
                #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
                #for persona in personas:
                ws.cell(row=cont,column=2).value = data['surname']
                ws.cell(row=cont,column=4).value = data['name'] 
                ws.cell(row=cont,column=5).value = data['edad']
                ws.cell(row=cont,column=7).value = data['nacionalidad']
                ws.cell(row=cont,column=8).value = data['formacion'] 
                ws.cell(row=cont,column=9).value = data['especifica']
                ws.cell(row=cont,column=11).value = 'Disponibilidad media jornada: ' + data['media_jornada'] + ' Disponibilidad entre semana: ' + data['entre_semana'] + 'Disponibilidad fines de semana: ' + data['fines_semana'] 
                ws.cell(row=cont,column=12).value = data['ciudad']
                ws.cell(row=cont,column=14).value = data['vehiculo'] 
                ws.cell(row=cont,column=15).value = data['email']
                ws.cell(row=cont,column=16).value = data['phone'] 
                ws.cell(row=cont,column=19).value = data['dni']
                ws.cell(row=cont,column=20).value = data['duracion']
                ws.cell(row=cont,column=21).value = data['desplazamiento']
                ws.cell(row=cont,column=22).value = data['solo_localidad']
                ws.cell(row=cont,column=23).value = data['provincia']
                ws.cell(row=cont,column=24).value = data['message']
                #cont = cont + 1
                #Establecemos el nombre del archivo
                nombre_archivo ="DatosTrabajador" + data['dni'] +".xlsx"
                wb.save(nombre_archivo)
             
                msg = EmailMessage('Bolsa de trabajo', 'Contacto recibido, No se ha adjuntado archivo, solo se adjunta la información en el excel', data['email'], ['info@manospararespirar.com'])
                msg.content_subtype = "html"  
                msg.attach_file(nombre_archivo)
                msg.send()
						  
			  
                print(data['email'])		
                return render(request, 'blog/mensaje.html', {})
 
            else:
                print('ERROR')
                return render(request, 'blog/mensaje.html', {})
 
    else:
        form = WorkForm()
        return render(request, 'blog/trabaja.html', {'form': form})

def indexeng(request):
    if request.method == 'POST':

        form = ContactForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            try:
                send_mail(
                    'Cliente','Asunto: '+ data['subject'] + 
                    ' El cliente: ' + data['name'] + ' Con email: ' + data['email'] + ' Mensaje: ' + data['message'] + ' Telefono: ' + data['phone'],
                    data['email'], #FROM
                    ['info@manospararespirar.com'],
                    fail_silently=False,
                )
			    
                print(data['email'])
				
                send_mail(
                    'Contact received', 'We have received your information. We will get in touch with you as soon as possible. \nThank you very much for trusting us. \nBest wishes form the whole team',
                    data['email'], #FROM
                    [data['email']],
                    fail_silently=False,
                )
				
                return render(request, 'blog/mensaje.html', {})
            except Exception as e:
                trace_back = traceback.format_exc()
                message = str(e)+ " " + str(trace_back)
                print (message)
                return render(request, 'blog/mensaje.html', {})
				
    else:
        return render(request, 'blog/indexmanos_eng.html', {})
 
def indexfr(request):
    if request.method == 'POST':

        form = ContactForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            send_mail(
                'Cliente','Asunto: '+ data['subject'] + 
                ' El cliente: ' + data['name'] + ' Con email: ' + data['email'] + ' Mensaje: ' + data['message'] + ' Telefono: ' + data['phone'],
                data['email'], #FROM
                ['info@manospararespirar.com'],
                fail_silently=False,
            )
      
            print(data['email'])
            return render(request, 'blog/mensaje.html', {})
    else:
        return render(request, 'blog/indexmanos_fr.html', {})
 
	
def indexmanos(request):
    if request.method == 'POST':

        form = ContactForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            try:
                #send_mail(
                #    'Cliente','Asunto: '+ data['subject'] + 
                #    ' El cliente: ' + data['name'] + ' Con email: ' + data['email'] + ' Mensaje: ' + data['message'] + ' Telefono: ' + data['phone'],
                #    data['email'], #FROM
                #    ['info@manospararespirar.com'],
                #    fail_silently=False,
                #)	
			    
                print(data['email'])
				
                send_mail(
                    'Información recibida', 'Hemos recibido su solicitud. Nos pondremos en contacto con usted lo antes posible. \n\nMuchas gracias por confiar en nosotros. \n\nUn saludo de todo el equipo',
                    data['email'], #FROM
                    [data['email']],
                    fail_silently=False,
                )
				
                return render(request, 'blog/mensaje.html', {})
            except Exception as e:
                trace_back = traceback.format_exc()
                message = str(e)+ " " + str(trace_back)
                print (message)
                return render(request, 'blog/mensaje.html', {})
    else:
        return render(request, 'blog/indexmanos.html', {})
 
	
def registro(request):
    if request.method == "POST":
        form = UsuarioForm(request.POST)
        if form.is_valid():
            usuario = form.save(commit=False)
            # usuario.nombre = form.fields.nombre
            # usuario.apellidos = form.apellidos
            # usuario.telefono_1 = form.telefono_1
            # usuario.telefono_2 = form.telefono_2
            # usuario.provincia = form.provincia
            # usuario.localidad = form.localidad
            # usuario.direccion = form.direccion
            # usuario.telefono_2 = form.telefono_2
            # usuario.codigo_postal = form.codigo_postal
            # usuario.carne_conducir = form.carne_conducir
            # usuario.sexo = form.sexo
            # usuario.cuidador = form.cuidador
            # usuario.administrador = False
            # usuario.fecha_nacimiento = form.fecha_nacimiento
			
            # email = form.email
            # email = email.lower().strip() # Hopefully reduces junk to ""
            # if email != "": # If it's not blank
                # if not email_re.match(email): # If it's not an email address
                    # raise ValidationError(u'%s is not an email address, dummy!' % email)
            # if email == "":
                # email = None
            # usuario.email = email
            usuario.save()

            return redirect('blog/post_list.html', pk=usuario.pk)
    else:
        form = UsuarioForm()
    return render(request, 'blog/registro.html', {'form': form})